#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pre_run_check.py

在執行 runner 前檢查 company_details.json、company_details.xlsx 與 hits.json
的資料筆數是否一致，必要時自動更新，最後再呼叫 runner.main。
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Dict

from openpyxl import load_workbook

import simple_logger as log
from company_details_builder import build_and_save
from export_company_details import main as export_excel
from persistence import BASE_DIR, HITS_PATH

COMPANY_DETAILS_JSON = BASE_DIR / "company_details.json"
COMPANY_DETAILS_XLSX = BASE_DIR / "company_details.xlsx"


def _pair_count(d: Dict) -> int:
    """計算 ban/year pair 數量。"""
    return sum(len(v) for v in d.values() if isinstance(v, dict))


def _load_json(path: Path) -> Dict:
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _count_json_pairs(path: Path) -> int:
    return _pair_count(_load_json(path))


def _count_excel_rows(path: Path) -> int:
    if not path.exists():
        return 0
    wb = load_workbook(path)
    ws = wb.active
    count = 0
    for row in range(2, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1).value
        c2 = ws.cell(row=row, column=2).value
        if c1 is not None and c2 is not None:
            count += 1
    return count


def run_checks() -> None:
    while True:
        json_cnt = _count_json_pairs(COMPANY_DETAILS_JSON)
        xlsx_cnt = _count_excel_rows(COMPANY_DETAILS_XLSX)
        if json_cnt > xlsx_cnt:
            log.warn("company_details.xlsx 落後，準備自動更新…")
            export_excel()
            continue
        if xlsx_cnt > json_cnt:
            log.warn("company_details.xlsx 筆數大於 company_details.json，請確認。")
            if input("確認沒問題？(y/N): ").strip().lower() != "y":
                sys.exit(1)
        hits_cnt = _count_json_pairs(HITS_PATH)
        json_cnt = _count_json_pairs(COMPANY_DETAILS_JSON)
        if hits_cnt > json_cnt:
            log.warn("company_details.json 落後，準備自動更新…")
            build_and_save(
                input_path=str(HITS_PATH),
                output_path=str(COMPANY_DETAILS_JSON),
            )
            continue
        if json_cnt > hits_cnt:
            log.warn("company_details.json 筆數大於 hits.json，請確認。")
            if input("確認沒問題？(y/N): ").strip().lower() != "y":
                sys.exit(1)
        break
    log.success("驗證資料ok...已更新資料...")


def main() -> None:
    run_checks()
    import runner  # 延後匯入以避免參數互動提早發生
    runner.main()


if __name__ == "__main__":
    main()
