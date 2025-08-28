#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pre_run_check.py

在執行 runner 前檢查 company_details.json、company_details.xlsx 與 hits.json
的資料筆數是否一致，必要時自動更新，最後再呼叫 runner.main。
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Dict

from openpyxl import load_workbook

import fbfh_trade.logger as log
from fbfh_trade.company.builder import build_and_save
from fbfh_trade.company.exporter import main as export_excel
from fbfh_trade.persistence import BASE_DIR, HITS_PATH

COMPANY_DETAILS_JSON = BASE_DIR / "company_details.json"
COMPANY_DETAILS_XLSX = BASE_DIR / "company_details.xlsx"


def _pair_count(d: Dict) -> int:
    """計算 ban/year pair 數量。"""
    return sum(len(v) for v in d.values() if isinstance(v, dict))


def _load_json(path: Path) -> Dict:
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _count_json_pairs(path: Path) -> int:
    return _pair_count(_load_json(path))


def _count_excel_rows(path: Path) -> int:
    """
    僅計算資料列數（自第 2 列起算，第一列預設為標題列）。
    以第 1、2 欄皆非 None 視為有效資料列，避免因格式化/空白列被誤計。
    """
    if not path.exists():
        return 0
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return 0

    # 明確自第 2 列（跳過標題列）開始統計
    start_row = 2
    if ws.max_row < start_row:
        return 0

    count = 0
    for row in range(start_row, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1).value
        c2 = ws.cell(row=row, column=2).value
        if c1 is not None and c2 is not None:
            count += 1
    return count


def run_checks() -> None:
    while True:
        # 1) 對齊 company_details.json 與 company_details.xlsx
        json_cnt = _count_json_pairs(COMPANY_DETAILS_JSON)
        xlsx_cnt = _count_excel_rows(COMPANY_DETAILS_XLSX)

        if json_cnt > xlsx_cnt:
            log.warn("company_details.xlsx 落後，準備自動更新…")
            prev = xlsx_cnt
            export_excel()
            new_cnt = _count_excel_rows(COMPANY_DETAILS_XLSX)

            # 若沒有前進，避免無限迴圈
            if new_cnt <= prev:
                log.warn("更新後的 company_details.xlsx 筆數未增加，停止以避免重複執行。")
                sys.exit(1)
            # 確認前進後，回到迴圈重新計算
            continue

        if xlsx_cnt > json_cnt:
            log.warn("company_details.xlsx 筆數大於 company_details.json，請確認。")
            if input("確認沒問題？(y/N): ").strip().lower() != "y":
                sys.exit(1)

        # 2) 對齊 hits.json 與 company_details.json
        hits_cnt = _count_json_pairs(HITS_PATH)
        json_cnt = _count_json_pairs(COMPANY_DETAILS_JSON)

        if hits_cnt > json_cnt:
            log.warn("company_details.json 落後，準備自動更新…")
            prev = json_cnt
            build_and_save(
                input_path=str(HITS_PATH),
                output_path=str(COMPANY_DETAILS_JSON),
            )
            new_cnt = _count_json_pairs(COMPANY_DETAILS_JSON)

            # 若沒有前進，避免無限迴圈
            if new_cnt <= prev:
                log.warn("更新後的 company_details.json 筆數未增加，停止以避免重複執行。")
                sys.exit(1)
            # 確認前進後，回到迴圈重新計算
            continue

        if json_cnt > hits_cnt:
            log.warn("company_details.json 筆數大於 hits.json，請確認。")
            if input("確認沒問題？(y/N): ").strip().lower() != "y":
                sys.exit(1)

        # 三方數量一致或已確認無誤，即可離開
        break

    log.success("驗證資料ok...已更新資料...")


def main() -> None:
    run_checks()
    # 延後匯入以避免參數互動提早發生
    import runner
    runner.main()


if __name__ == "__main__":
    main()

