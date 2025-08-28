#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""建立 company_details.json 後，接著輸出 company_details.xlsx（皆放在 exe 同目錄）。"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Optional, Tuple

from openpyxl import load_workbook

import fbfh_trade.logger as log
from fbfh_trade.company.builder import build_and_save
from fbfh_trade.company.exporter import main as export_excel
from fbfh_trade.persistence import BASE_DIR  # 關鍵：exe 同目錄

INPUT_JSON_NAME = "hits.json"
OUTPUT_JSON_NAME = "company_details.json"
OUTPUT_XLSX_NAME = "company_details.xlsx"


def _load_existing(path: Path) -> Tuple[Dict[str, Dict[str, Any]], Optional[Tuple[str, str]]]:
    if not path.exists():
        return {}, None
    with path.open("r", encoding="utf-8") as f:
        data: Dict[str, Dict[str, Any]] = json.load(f)

    last_ban: Optional[str] = None
    last_year: Optional[str] = None
    for ban_no, years in data.items():
        if not isinstance(years, dict):
            continue
        last_ban = ban_no
        for year in years.keys():
            last_year = year
    if last_ban and last_year:
        return data, (last_ban, last_year)
    return data, None


def _excel_last_pair(path: Path) -> Optional[Tuple[str, str]]:
    if not path.exists():
        return None
    wb = load_workbook(path)
    ws = wb.active
    for row in range(ws.max_row, 0, -1):
        tax = ws.cell(row=row, column=1).value
        year = ws.cell(row=row, column=6).value
        if tax is not None:
            return str(tax), str(year)
    return None


def _slice_hits_after(path: Path, last_pair: Optional[Tuple[str, str]]) -> Dict[str, Dict[str, Any]]:
    with path.open("r", encoding="utf-8") as f:
        hits = json.load(f)

    if not last_pair:
        return hits

    started = False
    sliced: Dict[str, Dict[str, Any]] = {}
    for ban_no, years in hits.items():
        for year, meta in years.items():
            if not started:
                if ban_no == last_pair[0] and year == last_pair[1]:
                    started = True
                continue
            sliced.setdefault(ban_no, {})[year] = meta

    if not started:
        log.warn("在 hits.json 中找不到最後處理的紀錄，將從頭開始處理。")
        return hits
    return sliced


def run_pipeline() -> None:
    """先建置 JSON，再匯出 Excel。"""
    log.PRINT_DEBUG = 1
    log.PRINT_INFO = 1
    log.PRINT_WARN = 1
    log.PRINT_ERROR = 1
    log.PRINT_SUCCESS = 1

    base = BASE_DIR
    input_path = base / INPUT_JSON_NAME
    output_path = base / OUTPUT_JSON_NAME
    excel_path = base / OUTPUT_XLSX_NAME

    existing, last_pair = _load_existing(output_path)
    if last_pair:
        excel_last = _excel_last_pair(excel_path)
        if excel_last != last_pair:
            print("發現 company_details.json 與 company_details.xlsx 不同步，準備自動更新")
            export_excel()

    hits_to_process = _slice_hits_after(input_path, last_pair)
    if not hits_to_process:
        log.info("沒有新的資料需要處理。")
        return

    tmp_hits = base / "hits.partial.json"
    with tmp_hits.open("w", encoding="utf-8") as f:
        json.dump(hits_to_process, f, ensure_ascii=False, indent=2)

    tmp_output = base / "company_details.partial.json"
    result = build_and_save(input_path=str(tmp_hits), output_path=str(tmp_output), timeout=10)

    if tmp_hits.exists():
        tmp_hits.unlink()
    if tmp_output.exists():
        tmp_output.unlink()

    for ban_no, years in result.items():
        existing.setdefault(ban_no, {}).update(years)

    with output_path.open("w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    export_excel()


if __name__ == "__main__":
    run_pipeline()
